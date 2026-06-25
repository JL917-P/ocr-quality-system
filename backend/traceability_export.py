"""Exportación de trazabilidad por lote (y producto) a Excel."""
from __future__ import annotations

import io
import re
import sqlite3
from collections import defaultdict
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from constancia_utils import (
    dedupe_constancia_rows,
    find_items_json_for_constancia,
    item_display_expiration,
    item_display_lot,
    item_display_product,
    item_display_production,
    parse_items_json,
    sort_constancia_rows_by_issue_date,
)

HEADERS = (
    "N°",
    "Fecha envío",
    "Cliente",
    "Producto",
    "Lote",
    "F. Producción",
    "F. Vencimiento",
    "Cantidad",
)

HEADER_FILL = PatternFill("solid", fgColor="B4C6E7")
THIN = Side(style="thin")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def normalize_key(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"\s+", " ", text)
    return text


def normalize_lot_key(value: Any) -> str:
    return normalize_key(value).replace(" ", "").replace("-", "")


def format_fecha_envio(fecha: str) -> str:
    raw = (fecha or "").strip()
    if not raw:
        return ""
    if re.match(r"^\d{4}-\d{2}-\d{2}", raw):
        try:
            dt = datetime.strptime(raw[:10], "%Y-%m-%d")
            return dt.strftime("%d / %m / %Y")
        except ValueError:
            pass
    slash = re.match(r"^(\d{2})\s*/\s*(\d{2})\s*/\s*(\d{4})", raw)
    if slash:
        dd, mm, yyyy = slash.groups()
        return f"{dd} / {mm} / {yyyy}"
    return raw


def parse_fecha_sort_key(fecha: str) -> float:
    raw = (fecha or "").strip()
    if not raw:
        return 0.0
    if re.match(r"^\d{4}-\d{2}-\d{2}", raw):
        try:
            return datetime.strptime(raw[:10], "%Y-%m-%d").timestamp()
        except ValueError:
            return 0.0
    slash = re.match(r"^(\d{2})\s*/\s*(\d{2})\s*/\s*(\d{4})", raw)
    if slash:
        dd, mm, yyyy = slash.groups()
        try:
            return datetime(int(yyyy), int(mm), int(dd)).timestamp()
        except ValueError:
            return 0.0
    try:
        return datetime.fromisoformat(raw).timestamp()
    except ValueError:
        return 0.0


def load_trace_rows(conn: sqlite3.Connection) -> list[dict[str, Any]]:
    rows = conn.execute(
        """
        SELECT id, number, issue_date, client_name, transport_plate, fumigacion, calidad, status, items_json, created_at
        FROM constancias
        """
    ).fetchall()
    rows = sort_constancia_rows_by_issue_date(dedupe_constancia_rows(list(rows)))
    trace_rows: list[dict[str, Any]] = []
    for row in rows:
        items = parse_items_json(row[8])
        if not items:
            alt_json = find_items_json_for_constancia(conn, row[1] or "", row[3] or "", exclude_id=row[0])
            if alt_json:
                items = parse_items_json(alt_json)
        fecha = row[2] or ""
        cliente = row[3] or ""
        for item in items:
            lot = item_display_lot(item)
            product = item_display_product(item)
            if not lot and not product:
                continue
            trace_rows.append(
                {
                    "fecha": fecha,
                    "cliente": cliente,
                    "product": product,
                    "lot": lot,
                    "production": item_display_production(item),
                    "expiration": item_display_expiration(item),
                    "quantity": item.get("quantity") if item.get("quantity") is not None else "",
                }
            )
    return trace_rows


def filter_trace_rows(
    trace_rows: list[dict[str, Any]],
    lote: str,
    producto: str | None = None,
) -> list[dict[str, Any]]:
    lot_key = normalize_lot_key(lote)
    if not lot_key:
        return []
    filtered = [r for r in trace_rows if normalize_lot_key(r.get("lot")) == lot_key]
    if producto and producto.strip():
        product_key = normalize_key(producto)
        filtered = [r for r in filtered if normalize_key(r.get("product")) == product_key]
    filtered.sort(key=lambda r: parse_fecha_sort_key(r.get("fecha", "")), reverse=True)
    return filtered


def products_for_lot(trace_rows: list[dict[str, Any]], lote: str) -> list[str]:
    filtered = filter_trace_rows(trace_rows, lote)
    products = sorted({str(r.get("product") or "").strip() for r in filtered if r.get("product")}, key=str.lower)
    return [p for p in products if p]


def group_trace_rows_by_product(rows: list[dict[str, Any]]) -> list[tuple[str, list[dict[str, Any]]]]:
    buckets: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        buckets[str(row.get("product") or "").strip() or "(Sin producto)"].append(row)
    return sorted(buckets.items(), key=lambda item: item[0].lower())


def _style_header_row(ws, row_idx: int) -> None:
    for col_idx, title in enumerate(HEADERS, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=title)
        cell.font = Font(bold=True, size=11)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = CELL_BORDER


def _write_data_rows(ws, start_row: int, rows: list[dict[str, Any]]) -> int:
    row_idx = start_row
    for idx, row in enumerate(rows, start=1):
        values = (
            idx,
            format_fecha_envio(row.get("fecha", "")),
            row.get("cliente", ""),
            row.get("product", ""),
            row.get("lot", ""),
            row.get("production", ""),
            row.get("expiration", ""),
            row.get("quantity", ""),
        )
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = CELL_BORDER
            cell.alignment = Alignment(
                horizontal="center" if col_idx in (1, 2, 5, 6, 7, 8) else "left",
                vertical="center",
                wrap_text=col_idx in (3, 4),
            )
        row_idx += 1
    return row_idx


def _autosize_columns(ws, max_row: int) -> None:
    widths = [6, 14, 34, 42, 12, 14, 14, 10]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def build_traceability_workbook(
    trace_rows: list[dict[str, Any]],
    lote: str,
    producto: str | None = None,
) -> bytes:
    filtered = filter_trace_rows(trace_rows, lote, producto)
    if not filtered:
        raise ValueError("No hay registros para ese lote y producto.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Trazabilidad"

    groups: list[tuple[str, list[dict[str, Any]]]]
    if producto and producto.strip():
        groups = [(producto.strip(), filtered)]
    else:
        groups = group_trace_rows_by_product(filtered)
        if len(groups) > 1:
            pass  # varias tablas, una por producto con el mismo lote

    row_idx = 1
    for group_index, (product_name, group_rows) in enumerate(groups):
        if group_index > 0:
            row_idx += 1
        if len(groups) > 1:
            title = ws.cell(row=row_idx, column=1, value=f"Producto: {product_name}")
            title.font = Font(bold=True, size=11)
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(HEADERS))
            row_idx += 1
        _style_header_row(ws, row_idx)
        row_idx = _write_data_rows(ws, row_idx + 1, group_rows)

    _autosize_columns(ws, row_idx)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_traceability_workbook_batch(
    trace_rows: list[dict[str, Any]],
    selections: list[dict[str, str]],
) -> bytes:
    if not selections:
        raise ValueError("Indique al menos un lote y producto.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Trazabilidad"
    row_idx = 1
    wrote_any = False

    for sel_index, sel in enumerate(selections):
        lote = str(sel.get("lote") or "").strip()
        producto = str(sel.get("producto") or "").strip()
        if not lote or not producto:
            continue
        filtered = filter_trace_rows(trace_rows, lote, producto)
        if not filtered:
            continue
        if wrote_any:
            row_idx += 1
        title = ws.cell(row=row_idx, column=1, value=f"Lote: {lote}  |  Producto: {producto}")
        title.font = Font(bold=True, size=11)
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(HEADERS))
        row_idx += 1
        _style_header_row(ws, row_idx)
        row_idx = _write_data_rows(ws, row_idx + 1, filtered)
        wrote_any = True

    if not wrote_any:
        raise ValueError("No hay registros para las selecciones indicadas.")

    _autosize_columns(ws, row_idx)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def unique_lots(trace_rows: list[dict[str, Any]]) -> list[str]:
    lots = sorted({str(r.get("lot") or "").strip() for r in trace_rows if r.get("lot")}, key=str.lower)
    return [lot for lot in lots if lot]
